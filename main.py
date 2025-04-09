import requests
import os
import tempfile
import psycopg2
import openpyxl
from dotenv import load_dotenv
from datetime import datetime
from psycopg2.extras import execute_values


def create_table(conn):
    """Create FTS table if it doesn't exist"""
    cursor = conn.cursor()

    # Columns from the Excel file comment in the original code
    cursor.execute(
        """
    CREATE TABLE IF NOT EXISTS fts_data (
        id SERIAL PRIMARY KEY,
        year INTEGER,
        budget TEXT,
        reference_legal_commitment TEXT,
        reference_budget TEXT,
        beneficiary_name TEXT,
        beneficiary_vat TEXT,
        not_for_profit BOOLEAN,
        non_governmental BOOLEAN,
        coordinator BOOLEAN,
        address TEXT,
        city TEXT,
        postal_code TEXT,
        beneficiary_country TEXT,
        nuts2 TEXT,
        geographical_zone TEXT,
        action_location TEXT,
        beneficiary_contracted_amount NUMERIC,
        beneficiary_estimated_contracted_amount NUMERIC,
        beneficiary_estimated_consumed_amount NUMERIC,
        commitment_contracted_amount NUMERIC,
        additional_reduced_amount NUMERIC,
        commitment_total_amount NUMERIC,
        commitment_consumed_amount NUMERIC,
        source_estimated_detailed_amount TEXT,
        expense_type TEXT,
        subject_grant_contract TEXT,
        responsible_department TEXT,
        budget_line_number TEXT,
        budget_line_name TEXT,
        programme_name TEXT,
        funding_type TEXT,
        beneficiary_group_code TEXT,
        beneficiary_type TEXT,
        project_start_date DATE,
        project_end_date DATE,
        type_of_contract TEXT,
        management_type TEXT,
        benefiting_country TEXT
    )
    """
    )

    conn.commit()


def clean_value(value, column_type=None):
    """Clean and normalize a single value"""
    if value is None or value == "":
        return None

    # Convert Yes/No to boolean
    if column_type == "boolean":
        if value == "Yes":
            return True
        elif value == "No":
            return False
        return None

    # Convert date values
    if column_type == "date":
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        elif value == "-" or not value:
            return None

    # Convert numeric values
    if column_type == "numeric":
        if isinstance(value, (int, float)):
            return value
        elif isinstance(value, str):
            try:
                return float(value.replace(",", ""))
            except (ValueError, AttributeError):
                return None
        return None

    # Return the value as-is for other types
    return value


def get_column_mapping():
    """Get mapping between Excel columns and database columns"""
    return {
        "Year": "year",
        "Budget": "budget",
        "Reference of the Legal Commitment (LC)": "reference_legal_commitment",
        "Reference (Budget)": "reference_budget",
        "Name of beneficiary": "beneficiary_name",
        "VAT number of beneficiary": "beneficiary_vat",
        "Not-for-profit organisation (NFPO)": "not_for_profit",
        "Non-governmental organisation (NGO)": "non_governmental",
        "Coordinator": "coordinator",
        "Address": "address",
        "City": "city",
        "Postal code": "postal_code",
        "Beneficiary country": "beneficiary_country",
        "NUTS2": "nuts2",
        "Geographical Zone": "geographical_zone",
        "Action location": "action_location",
        "Beneficiary's contracted amount (EUR)": "beneficiary_contracted_amount",
        "Beneficiary's estimated contracted amount (EUR)": "beneficiary_estimated_contracted_amount",
        "Beneficiary's estimated consumed amount (EUR)": "beneficiary_estimated_consumed_amount",
        "Commitment contracted amount (EUR) (A)": "commitment_contracted_amount",
        "Additional/Reduced amount (EUR) (B)": "additional_reduced_amount",
        "Commitment  total amount (EUR) (A+B)": "commitment_total_amount",
        "Commitment consumed amount (EUR)": "commitment_consumed_amount",
        "Source of (estimated) detailed amount": "source_estimated_detailed_amount",
        "Expense type": "expense_type",
        "Subject of grant or contract": "subject_grant_contract",
        "Responsible department": "responsible_department",
        "Budget line number": "budget_line_number",
        "Budget line name": "budget_line_name",
        "Programme name": "programme_name",
        "Funding type": "funding_type",
        "Beneficiary Group Code": "beneficiary_group_code",
        "Beneficiary type": "beneficiary_type",
        "Project start date": "project_start_date",
        "Project end date": "project_end_date",
        "Type of contract*": "type_of_contract",
        "Management type": "management_type",
        "Benefiting country": "benefiting_country",
    }


def get_column_types():
    """Get column types for data cleaning"""
    return {
        "Not-for-profit organisation (NFPO)": "boolean",
        "Non-governmental organisation (NGO)": "boolean",
        "Coordinator": "boolean",
        "Project start date": "date",
        "Project end date": "date",
        "Beneficiary's contracted amount (EUR)": "numeric",
        "Beneficiary's estimated contracted amount (EUR)": "numeric",
        "Beneficiary's estimated consumed amount (EUR)": "numeric",
        "Commitment contracted amount (EUR) (A)": "numeric",
        "Additional/Reduced amount (EUR) (B)": "numeric",
        "Commitment  total amount (EUR) (A+B)": "numeric",
        "Commitment consumed amount (EUR)": "numeric",
    }


def process_excel_data(file_path, year):
    """Process Excel data using openpyxl for better performance"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # Get column headers from first row
    headers = []
    for cell in next(ws.iter_rows()):
        headers.append(cell.value)

    print(f"Columns for year {year}:")
    print(headers)

    # Get column mapping and types
    column_mapping = get_column_mapping()
    column_types = get_column_types()

    # Determine which columns exist in both Excel and database
    excel_to_db_columns = {}
    db_columns = []

    for idx, header in enumerate(headers):
        if header in column_mapping:
            excel_to_db_columns[idx] = column_mapping[header]
            db_columns.append(column_mapping[header])

    # Prepare column types for processing
    excel_column_types = {}
    for idx, header in enumerate(headers):
        if header in column_types:
            excel_column_types[idx] = column_types[header]

    # Process data rows
    rows = []
    batch_size = 5000  # Larger batch size for better performance
    row_count = 0

    # Skip header row
    first_row = True

    for row in ws.iter_rows():
        if first_row:
            first_row = False
            continue

        row_values = []
        for idx, db_col in excel_to_db_columns.items():
            value = row[idx].value
            column_type = excel_column_types.get(idx)
            cleaned_value = clean_value(value, column_type)
            row_values.append(cleaned_value)

        rows.append(tuple(row_values))
        row_count += 1

        # When batch size is reached, yield the batch
        if len(rows) >= batch_size:
            yield db_columns, rows
            rows = []

    # Yield any remaining rows
    if rows:
        yield db_columns, rows

    wb.close()


def insert_data_batch(conn, db_columns, rows):
    """Insert data into the database using faster execute_values method"""
    cursor = conn.cursor()

    # Build the INSERT query
    columns_str = ", ".join(db_columns)

    # Use execute_values which handles the placeholders internally
    query = f"INSERT INTO fts_data ({columns_str}) VALUES %s"

    # Use the faster psycopg2.extras.execute_values method
    execute_values(cursor, query, rows, page_size=1000)
    conn.commit()


def insert_data(conn, file_path, year):
    """Process and insert Excel data into database"""
    total_rows = 0
    batch_count = 0

    for db_columns, rows in process_excel_data(file_path, year):
        batch_count += 1
        total_rows += len(rows)
        insert_data_batch(conn, db_columns, rows)
        print(f"Inserted {len(rows)} records for year {year} (batch {batch_count})")

    print(f"Total {total_rows} records inserted for year {year}")


def connect_to_database():
    """Connect to the PostgreSQL database using credentials from .env"""
    load_dotenv()

    db_host = os.getenv("DB_HOST", "localhost")
    db_port = os.getenv("DB_PORT", "5432")
    db_name = os.getenv("DB_NAME")
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")

    if not all([db_name, db_user, db_password]):
        raise ValueError("Database credentials missing from .env file")

    conn = psycopg2.connect(
        host=db_host, port=db_port, dbname=db_name, user=db_user, password=db_password
    )

    return conn


def main():
    headers = {
        "User-Agent": "Mozilla/5.0 (X11; Linux x86_64; rv:138.0) Gecko/20100101 Firefox/138.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        # 'Accept-Encoding': 'gzip, deflate, br, zstd',
        "Connection": "keep-alive",
        "Referer": "https://ec.europa.eu/budget/financial-transparency-system/help.html",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-User": "?1",
        "Priority": "u=0, i",
    }

    current_year = datetime.now().year

    try:
        # Connect to the PostgreSQL database
        conn = connect_to_database()

        # Create the table if it doesn't exist
        create_table(conn)

        # Create temporary directory
        with tempfile.TemporaryDirectory() as temp_dir:
            for year in range(2007, current_year + 1):
                url = f"https://ec.europa.eu/budget/financial-transparency-system/download/{year}_FTS_dataset_en.xlsx"
                print(f"Downloading data for year {year}...")

                try:
                    response = requests.get(url, headers=headers)

                    if response.status_code == 200:
                        print(f"Status code: {response.status_code} - Success")

                        # Save file to temporary location
                        temp_file_path = os.path.join(
                            temp_dir, f"{year}_FTS_dataset_en.xlsx"
                        )
                        with open(temp_file_path, "wb") as f:
                            f.write(response.content)

                        # Process and insert data using the refactored approach
                        insert_data(conn, temp_file_path, year)

                        print(f"Successfully processed data for year {year}")
                        print("-" * 50)

                    else:
                        print(
                            f"Status code: {response.status_code} - Failed to download"
                        )

                except Exception as e:
                    print(f"Error processing year {year}: {e}")

    except Exception as e:
        print(f"Database error: {e}")

    finally:
        if "conn" in locals() and conn:
            conn.close()
            print("Database connection closed")


if __name__ == "__main__":
    main()
